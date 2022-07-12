#实现excel表格按行数分拆的功能
from tkinter.messagebox import showinfo

import openpyxl
import datetime
import time
import tkinter as tk
from tkinter import filedialog
window = tk.Tk()
window.title('歌词信息提取')
window.geometry('500x250')
window.config(background='#d6d6d6')

lie1 = 20
lie2 = 190
tk.Label(window, text="请输入单个文件分割数量：", bg="#d6d6d6").place(x=lie1, y=30)
chaifen_cout = tk.StringVar()  # 文件输入路径变量
tk.Label(window, text="请选择要执行的文件：", bg="#d6d6d6").place(x=lie1, y=70)
var_name = tk.StringVar()  # 文件输入路径变量
tk.Label(window, text="请选择拆分文件后存储位置：",bg = "#d6d6d6").place(x=lie1, y=110)
var_name2 = tk.StringVar()  # 文件夹输入路径变量

entry_name = tk.Entry(window, textvariable=chaifen_cout, width=20)
entry_name.place(x=lie2, y=30)
entry_name = tk.Entry(window, textvariable=var_name, width=30)
entry_name.place(x=lie2, y=70)
entry_name2 = tk.Entry(window, textvariable=var_name2, width=30)
entry_name2.place(x=lie2, y=110)

# 文件选择
def selectPath_file():
    path_ = filedialog.askopenfilename(filetypes=[("数据表", [".xlsx"])])
    print(path_)
    var_name.set(path_)
# 输入文件夹路径
def selectPath_dir():
    path_t = filedialog.askdirectory()
    var_name2.set(path_t)

def chaifen():
    limit = int(chaifen_cout.get())  # 得到文本输入框的值
    files = var_name.get()
    path3 = var_name2.get()
    wb = openpyxl.load_workbook(files)
    sheet = wb['Sheet1']
    nrows = sheet.max_row  # 最大行数
    ncols = sheet.max_column  # 最大列数
    sheets = nrows / limit
    print(type(nrows))
    print(type(limit))
    if not sheets.is_integer():  #如果不是整除则需要+1
        sheets = int(sheets) + 1
        for i in range(1, sheets + 1):
            wb = openpyxl.Workbook()
            sheett = wb['Sheet']
            # 写入第一行数据
            for n in range(1, ncols + 1):
                sheett.cell(row=1, column=n).value = sheet.cell(row=1, column=n).value
            # 写入范围内数据
            t = 2 + limit * (i - 1)
            num_index = 2
            for row_num in range(t, t + limit):
                for col_num in range(1, ncols + 1):
                    sheett.cell(row=num_index, column=col_num).value = sheet.cell(row=row_num,column=col_num).value
                num_index = num_index + 1
            wb.save(path3+"/{excelname}.xlsx".format(excelname=i))
    else:
        for i in range(1,int(sheets)+1):
            wb =openpyxl.Workbook()
            sheett = wb['Sheet']
            # 写入第一行数据
            for n in range(1,ncols+1):
                sheett.cell(row=1,column=n).value=sheet.cell(row=1,column=n).value
            # 写入范围内数据
            t=2+limit*(i-1)
            num_index=2
            for row_num in range(t,t+limit):
                for col_num in range(1,ncols+1) :
                    sheett.cell(row=num_index,column=col_num).value=sheet.cell(row=row_num,column=col_num).value
                num_index=num_index+1
            wb.save(path3+"/{excelname}.xlsx".format(excelname=i))
    showinfo('提示', '已完成数据拆分!')
    window.quit()


tk.Button(window, text="文件选择", command=selectPath_file,bg = '#d1d1d1').place(x=410, y=65)
tk.Button(window, text="存储路径选择", command=selectPath_dir,bg = '#d1d1d1').place(x=410, y=100)
tk.Button(window, text="运行", command=chaifen,width = 15,height = 2,bg = '#d1d1d1').place(x=200, y=150)
window.mainloop()  # 显示窗口
